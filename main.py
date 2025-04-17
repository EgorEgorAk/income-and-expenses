import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime
from hide_token import TOKEN  # Убедись, что файл hide_token.py содержит TOKEN = "ваш_токен"
import telebot
from telebot import types

bot = telebot.TeleBot(TOKEN)
user_data = {}

# Команда /start
@bot.message_handler(commands=['start'])
def send_welcome(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add(types.KeyboardButton("Записать в таблицу"))
    bot.send_message(message.chat.id, "Привет! Нажми кнопку ниже, чтобы начать.", reply_markup=markup)

# Команда /help
@bot.message_handler(commands=['help'])
def send_help(message):
    bot.send_message(message.chat.id, "Просто нажми 'Записать в таблицу' и следуй инструкциям 🙂")

# Обработка кнопки
@bot.message_handler(func=lambda msg: msg.text == "Записать в таблицу")
def start_input(message):
    chat_id = message.chat.id
    user_data[chat_id] = {}
    bot.send_message(chat_id, "Введите дату (например, 10.10.2025):")
    bot.register_next_step_handler(message, get_date)

def get_date(message):
    chat_id = message.chat.id
    user_data[chat_id]['date'] = message.text
    bot.send_message(chat_id, "Введите тип (доходы или расходы):")
    bot.register_next_step_handler(message, get_type)

def get_type(message):
    chat_id = message.chat.id
    user_data[chat_id]['type'] = message.text.strip().lower()
    bot.send_message(chat_id, "Введите категорию или описание:")
    bot.register_next_step_handler(message, get_category)

def get_category(message):
    chat_id = message.chat.id
    user_data[chat_id]['category'] = message.text
    bot.send_message(chat_id, "Введите сумму (только число):")
    bot.register_next_step_handler(message, get_amount)

def get_amount(message):
    chat_id = message.chat.id
    try:
        amount = int(message.text)
    except ValueError:
        bot.send_message(chat_id, "❗ Ошибка: сумма должна быть числом! Попробуй снова.")
        return

    user_data[chat_id]['amount'] = amount

    entries = [[
        user_data[chat_id]['date'],
        user_data[chat_id]['type'],
        user_data[chat_id]['category'],
        amount
    ]]

    income_expenses(entries, chat_id)
    bot.send_message(chat_id, "✅ Данные записаны! Файл отправлен.")

# Функция записи и отправки Excel
def income_expenses(data, chat_id, filename="money.xlsx"):
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Доходы и Расходы"
        ws['A1'] = datetime.now().strftime("Создано: %Y-%m-%d %H:%M:%S")
        ws.append([])
        ws.append(['Дата', 'Тип', 'Категория', 'Сумма'])
        for cell in ws[3]:
            cell.font = Font(bold=True)

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row_data in data:
        ws.append(row_data)
        row_idx = ws.max_row
        fill = green_fill if row_data[1] == "доходы" else red_fill
        for cell in ws[row_idx]:
            cell.fill = fill

    # Подсчёт итогов
    total_income = 0
    total_expense = 0

    for row in ws.iter_rows(min_row=4, min_col=2, max_col=4):
        type_cell = row[0].value
        amount_cell = row[2].value
        if isinstance(amount_cell, (int, float)):
            if type_cell == "доходы":
                total_income += amount_cell
            elif type_cell == "расходы":
                total_expense += amount_cell

    balance = total_income - total_expense

    # Удаляем старые итоги
    rows_to_delete = []
    for row in ws.iter_rows(min_row=4):
        value = str(row[0].value)
        if value.startswith("Итого") or value.startswith("Остаток"):
            rows_to_delete.append(row[0].row)
    for row_idx in reversed(rows_to_delete):
        ws.delete_rows(row_idx)

    # Удаление пустых строк
    for row in reversed(range(4, ws.max_row + 1)):
        if all(cell.value is None for cell in ws[row]):
            ws.delete_rows(row)

    # Записываем итоги
    ws.append(["Итого доходы:", "", "", total_income])
    ws.append(["Итого расходы:", "", "", total_expense])
    ws.append(["Остаток:", "", "", balance])
    for i in range(3):
        for cell in ws[ws.max_row - i]:
            cell.font = Font(bold=True)

    wb.save(filename)

    # 📤 Отправка файла пользователю
    with open(filename, 'rb') as f:
        bot.send_document(chat_id, f)

# Запуск бота
bot.polling(none_stop=True)